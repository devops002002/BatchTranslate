# -*- coding:utf-8 -*-
from datetime import datetime
import time

import requests

import Country
from openpyxl import Workbook

base_url = 'https://translate.google.com/translate_a/single' \
           '?client=webapp' \
           '&sl={0}' \
           '&tl={1}&hl=zh-CN&dt=at&dt=bd&dt=ex&dt=ld&dt=md&dt=qca&dt=rw&dt=rm&dt=ss&dt=t&ssel=5&tsel=4&kc=0&tk={3}' \
           '&q={2}'


def read_file_to_list(path):
    '''
    读取文本内容并存到list里
    :param path:文件路径
    :return: list
    '''
    with open(path) as fp:
        content = fp.readlines()
    list = [x.strip() for x in content]
    return list


def get_content_list(path):
    '''
    获取要翻译的list内容
    :param path: 文件路径
    :return: list
    '''
    content_list = read_file_to_list(path)
    len1 = len(content_list)
    c_content_list = []
    for i in range(0, len1 - 1):
        c_content_list.append(content_list[i])
    return c_content_list


def get_tk(path):
    '''
    获取tk的参数值,tk的参数值位于文本的最后一行
    :param path: 路径
    :return: tk的参数值
    '''
    content_list = read_file_to_list(path)
    len1 = len(content_list)
    return content_list[len1 - 1]


def get_translate_result(source_country, target_country, content_list, tk):
    '''
    :param source_country: 翻译前的语言
    :param target_country: 翻译后的语言
    :param content: 翻译的内容
    :return: 结果
    '''

    content = '\n'.join(str(e) for e in content_list)
    URL = base_url.format(source_country, target_country, content, tk)
    json_result = requests.get(url=URL)
    data = json_result.json()
    result = []
    print source_country, '-------------->', target_country
    for i in range(0, len(content_list)):
        result.append(data[0][i][0])
    return result


if __name__ == '__main__':
    content_list = get_content_list('content.txt')
    tk = get_tk('content.txt')

    wb = Workbook()
    ws = wb.active
    ws.title = 'translate'
    _date = datetime.now().strftime('%Y%m%d')

    country_list = Country.get_country_list()
    len_count = len(country_list)

    for i in range(0, len_count):
        index = chr(65 + i) + str(1)
        row = ws[index]
        row.value = country_list[i]

    for i in range(0, len_count):
        result = get_translate_result(Country.Country.ChinaSimple, country_list[i], content_list, tk)
        for j in range(0, len(result)):
            index = chr(65 + i) + str(2 + j)
            row = ws[index]
            row.value = result[j]

    wb.save('translate_google_' + _date + '.xlsx')
